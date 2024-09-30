#import openpyxl as OP
#dw = OP.load_workbook('Archivo.xlsx')
#df = dw.active
#df.title = 'Titulo'
#
#print('Total number of rows: '+str(df.#max_row)+'. And total number of #columns: '+str(df.max_column))

#import openpyxl as OP
#Tb = OP.load_workbook('Tablas_dinámicas.xlsx')
#Tf = Tb.active
#def Con_fi_co(): #Recuperar el numero de filas y columnas dentro de mi archivo xlsx
    #print(f'Número de filas:{Tf.max_row} \n Número de columnas: {Tf.max_column}')
#print(f'El valor de la celda A1 es : {Tf['B2'].value}')  
#Recupera el valor de la celda apuntada

'''Leer datos de varias celdas'''
#values = [Tf.cell(row=1, column=i).value for i in range(1, Tf.max_column)] #Obtener todos los datos de las celdas de una fila
#print(values)
#valores = [Tf.cell(row=i, column=2).value for i in range(2, Tf.max_row)] #Obtener todos los datos de las celdas de una columna
#print(valores)

# reading data from a range of cells (from column 1 to 6)

#my_list = list()
#
#for value in Tf.iter_rows(
#    min_row=1, max_row=11, min_col=1, max_col=6, 
#    values_only=True):
#    my_list.append(value)
#for item in my_list:
#    print(item)


'''Escribir en archivos Excel con Openpyxl'''
'''Dos Alternativas'''
#df['A2'] = 'Valores1'
#df.cell(row=2, column=1, value='Valores1') #Para bucles

'''Crear una nueva columna'''
#La nueva columna suma los valores todos los valores de la column 5: Puntajes
#total_sales = 0
#for i in range(2, (Tf.max_row)+1):
#    total_sales += ((Tf.cell(row=i, column=5).value))
#Tf.cell(row=2, column=6).value = total_sales
#Tb.save('Tablas_dinámicas.xlsx')

'''Añadir nuevas filas'''
#new_row = (1,'The Legend of Zelda',1986,'Action','Nintendo',3.74,0.93,1.69,0.14,6.51,6.5)

#ws.append(new_row) #append -> sirve para añadir valores
    
#wb.save('videogamesales.xlsx')

#values = [ws.cell(row=ws.max_row,column=i).value for i in range(1,ws.max_column+1)]
#print(values) #->Confirmas que los datos se han añadido

'''Borrar filas'''
#ws.delete_rows(ws.max_row, 1) #delete_rows ->Borramos una fila

'''Crear fórmulas de Excel con Openpyxl'''
#Tf['F1'] = 'Total'
#Tf['F2'] = '=SUM(E2:E141)'
#Tb.save('Tablas_dinámicas.xlsx')


'''Trabajar con hojas en Openpyxl'''
#print(ws.title) #-> imprimir el nombre de la hoja activa con la cual stamos trabajando
#ws.title = 'Video Game'

'''Crear una nueva hoja de calculo'''
#wb.sheetnames #sheetnames -> muestra todas las hojas de cálculo del archivo

#wb.create_sheet('Empty Sheet') #-> Crea una nueva hoja dentro del archivo


'''Eliminar una hoja de cálculo'''
#wb.remove(['Empty Sheet']) #remove -> Elimina la hoja que eligas
#print(wb.sheetnames)

'''Duplica una hoja'''
#wb.copy_worksheet(wb['Video Game']) #copy_worksheet -> Crear una copia de una hoja de cálculo existente

'''Añadir gráficos'''
#from openpyxl.chart import Reference, #BarChart
##valores
#
#values = Reference(Tf, min_row=Tf.#min_row, max_row=Tf.max_row, #min_col=5, max_col=5)
#
##categorias
#
#cats_values = Reference(Tf, min_row=Tf.#min_row, max_row=Tf.max_row, #min_col=4, max_col=4)
#
##Crear diagrama de barras
#Diagrama = BarChart()
#
#Diagrama.add_data(values, #titles_from_data=True) #Agregar #valores al diagrama
#Diagrama.set_categories(cats_values) ##Agregar categorias al diagrama
#
##Establecer titulos 
#Diagrama.title = 'Titulo'
#Diagrama.x_axis.title = 'Titulo X'
#Diagrama.y_axis.title = 'Titulo Y'
#
##Agregar el grafico a la hoja activa
#Tf.add_chart(Diagrama, 'G4')
#
#Tb.save('Tablas_dinámicas.xlsx')

#'''Formatear celdas con Openpyxl'''
#from openpyxl.styles import Font, colors, PatternFill, Border, Side
#
##Font -> Letra
##colors -> Color
##PatternFill -> Fondo de las celdas
##Border -> borde de la celda
##Side -> Tamaño en general
#Tf['F1'].font = Font(color='FF0000',bold=True, size=12)
#Tf['F1'].fill = PatternFill('solid', start_color='38e3ff')
#
#my_border = Side(border_style='thin', color='000000')
#Tf['F1'].border = Border(top=my_border, left=my_border,#right=my_border, bottom=my_border)
#Tb.save('Tablas_dinámicas.xlsx')

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import	PatternFill

#fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
#Tf.conditional_formatting.add('E2:E141', CellIsRule(operator='greaterThan', formula=[8], fill=fill))
#
#Tb.save('Tablas_dinámicas.xlsx')
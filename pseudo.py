import openpyxl as OP
from openpyxl.chart import BarChart, Reference
import json


dw = OP.load_workbook('Consolidado_de_ventas.xlsx')
#Con load_workbook cargamos el archivo en especifico



df = dw['ENERO']
# con dw['ENERO'] seleccionamos una hoja


Valores = Reference(df,min_row = 3 , max_row = 7, min_col = 1, max_col = 4)
#Escribimos los valores que van en y
Cat_Valores = Reference(df,min_row = 3 , max_row = 7, min_col = 1, max_col = 1)
#Escribimos los valores que van en x

Diagrama = BarChart()
#Creamos el diagrama
Diagrama.add_data(Valores, titles_from_data = True)
#Agregamos y en el diagrama
Diagrama.set_categories(Cat_Valores)
#Agregamos x en el diagrama

Diagrama.title = 'Titulo'
Diagrama.x_axis.title = 'Vendedora'
Diagrama.y_axis.title = 'Costo'
#Agregamos los titulos del diagrama

df.add_chart(Diagrama, 'G4')
#Especificamos en que parte de la hoja del archivo excel colocar
dw.save('Consolidado_de_ventas.xlsx')
#Cargarmos los cambios en la hoja.




#def Print_Conten_Hojas():
#    my_list2 = list()
#    for item in df:
#        ds = dw[f'{item}']
#        for valor in ds.iter_rows(min_row=3,max_row=7,min_col=1,max_col=4, values_only=True):
#            my_list2.append(valor)
#        for items in my_list2:
#            print(items)
#Print_Conten_Hojas()

#df = dw['FEBRERO']
#my_list = list()
#
#for value in df.iter_rows(
#    min_row=3, max_row=7, min_col=1, max_col=4, 
#    values_only=True):
#    my_list.append(value)
#for item in my_list:
#    print(item)
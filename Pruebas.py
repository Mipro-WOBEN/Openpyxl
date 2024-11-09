import openpyxl as opx

df = opx.load_workbook("Tablas_dinamicas.xlsx")
dw = df.active
values = [dw.cell(row=1, column=i).value for i in range(1, dw.max_column)]
valuesrow = [dw.cell(row=i, column=1).value for i in range(1, dw.max_row)]

my_list = list()

for value in dw.iter_rows(
    min_row=1, max_row=11, min_col=1, max_col=6, 
    values_only=True):
    my_list.append(value)
for item in my_list:
    print(item)
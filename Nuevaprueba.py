import pandas as pd
import openpyxl as op
from openpyxl.styles import Font,PatternFill, Border,Side

dw = op.load_workbook('Archivo.xlsx')
df = dw.active


Fondo_Celdas = PatternFill('solid', start_color='464ce6')
my_border = Side(color='040520', border_style='double')
Format_Letra = Font(color='56ff33', size=12, name='Oswald')
Lista_Alfabetica = ['A','B','C','D']
for i in Lista_Alfabetica:
    df[f'{i}1'].fill = Fondo_Celdas
    df[f'{i}1'].font = Format_Letra
    df[f'{i}1'].border = Border(left=my_border,right=my_border,top=my_border,bottom=my_border)

dw.save('Archivo.xlsx')



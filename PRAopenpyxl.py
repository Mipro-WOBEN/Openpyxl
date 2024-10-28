import pandas as pd
from openpyxl import load_workbook #Cargar un archivo excel.
from openpyxl.chart import BarChart, Reference #Realizar graficas en excel
from openpyxl.styles import Font
import string

def automatizar_excel(nombre_archivo):
    #1 leer archivos excel
    """Input sales_mes.xlsx/ Output report_mes.xlsx"""
    archivo_excel = pd.read_excel(nombre_archivo) #Referencia para #2

    #2Tablas Pivote
    Tabla_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
    mes_extension = input('Introduce el nombre con el cual quieres guardar el reporte: ')
    Tabla_pivote.to_excel(f'sales_{mes_extension}', startrow=4, startcol=0, sheet_name='Report')

    #cargar archivos con openpyxl

    wb = load_workbook(f'sales_{mes_extension}')
    pestana = wb['Report']

    mi_colum = wb.active.min_column
    max_colum = wb.active.max_column
    mi_fila = wb.active.min_row
    max_fila = wb.active.max_row

    #grafico

    barchart = BarChart() #Para crear Barras en excel
    data = Reference(pestana, min_col=mi_colum+1, max_col=max_colum,min_row=mi_fila , max_row=max_fila) #Incluir datos en el grafico
    categoria = Reference(pestana, min_col=mi_colum, max_col=mi_colum,min_row=mi_fila+1, max_row=max_fila) #Incluir categorias

    barchart.add_data(data, titles_from_data=True) #Añade datos al grafico
    barchart.set_categories(categoria)#Define las categorias del grafico

    pestana.add_chart(barchart, 'B2') #Determina en donde quiero el grafico
    barchart.title = 'Ventas' #Describe el nombre del grafico
    barchart.style = 5 #Aplica el estilo del  grafico


    #Aplicar formulas de excel
    abecedario = string.ascii_uppercase #Creamos una variable con el abecedario en mayuscula
    abecedario_excel = abecedario[0:max_colum] #Definimos una varible con el abecedario y un rango

    for i in abecedario_excel:
        if i!='A':
            pestana[f'{i}{max_fila+1}'] = f'=SUM({i}{mi_fila+1}:{i}{max_fila})' #Adición de la formula 
            #suma 
            pestana[f'{i}{max_fila+1}'].style = 'Currency' #estilo de moneda dolar
    pestana[f'{abecedario[0]}{max_fila+1}'] = 'Total'

    pestana['A1'] = 'Reporte'
    mes = mes_extension.split('.')[0]
    pestana['A2'] = mes
    pestana['A1'].font = Font('Arial', bold=True, size=20)
    pestana['A2'].font = Font('Arial', bold=True, size=12)

    wb.save(f'sales_{mes_extension}')
    return
automatizar_excel('supermarket_sales.xlsx')